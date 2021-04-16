import csv
import datetime
import copy
import openpyxl
# 格式2019/12/31 23:11

report_name = 'report.csv'
src_url = '原始报告/' + report_name
#year0 = datetime.datetime.now().strftime("%Y")
save_url = '处理后的报告/' +'处理后的报告.xlsx'
"""
初始化
"""
# 表单
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "精简数据"
ws2 = wb.create_sheet("365 Days")
ws3 = wb.create_sheet("52 Weeks")
ws4 = wb.create_sheet("12 Months")
# 字典
days_dict = {}

"""
表单1:"精简数据"
"""
with open(src_url, 'r', encoding='utf-8') as f:
    read = csv.DictReader(f)
    print("开始制作'精简数据'表单...\n")
    ws.append(["活动类别", "持续时间", "开始时间", "结束时间"])
    for row in read:
        if row["结束时间"] != None:
            comment_act = row["备注"]
            act = row["活动类别"]
            start_time = datetime.datetime.strptime(
                row["开始时间"], "%Y/%m/%d %H:%M:%S")
            over_time = datetime.datetime.strptime(
                row["结束时间"], "%Y/%m/%d %H:%M:%S")
            alt = over_time-start_time
            if start_time.day != over_time.day:
                zero_datetime = datetime.datetime(
                    over_time.year, over_time.month, over_time.day)
                alt01 = zero_datetime-start_time
                alt02 = over_time-zero_datetime
                row_list01 = [act, alt01, start_time,
                              zero_datetime, comment_act]
                row_list02 = [act, alt02, zero_datetime,
                              over_time, comment_act+"(昨天的延续)"]
                ws.append(row_list01)
                ws.append(row_list02)
            else:
                row_list = [act, alt, start_time, over_time, comment_act]
                ws.append(row_list)
    # 开始设置列宽
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 35
    ws.delete_rows(3)
    ws.delete_rows(ws.max_row-1)
    print("'精简数据'表单制作完成!lol\n")

"""
表单2:"365Days"
"""
activity_name = sorted(list(set([x.value for x in ws["A"][1:]])))
rq = sorted(list(set([x.value.date() for x in ws["C"][1:]])))
print("开始制作365Days表单\n")
# 先创建好一个字典容器
for act in activity_name:
    days_dict[act] = {}
    for d in rq:
        days_dict[act][d] = 0
# 给容器添加数据
for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    act = row[0].value
    alt = row[1].value
    start_date = row[2].value.date()
    st = round(alt.seconds/3600, 3)
    days_dict[act][start_date] += st
# 制作表头
head1 = copy.deepcopy(activity_name)
head1.insert(0, "日期")
ws2.append(head1)
ws3.append(head1)
ws4.append(head1)
# 将字典里的数据按顺序,一行行填写进表格2
for d in rq:
    l = [d]
    for act in activity_name:
        l.append(days_dict[act][d])
    ws2.append(l)
ws2.column_dimensions['A'].width = 12
ws2.delete_rows(2)
ws2.delete_rows(ws2.max_row)
print("365Days表单制作完成\n")

"""
表单3:"52 Weeks"
"""
# 制作容器
rq_365 = sorted(list(set([x.value for x in ws2["A"][1:]])))
zz = []
mm = []
for d in rq_365:
    if d.strftime("%W") not in zz:
        zz.append(d.strftime("%W"))
    if str(d.month) not in mm:
        mm.append(str(d.month))

for act in activity_name:
    days_dict[act]["week"] = {}
    days_dict[act]["month"] = {}
    for z in zz:
        days_dict[act]["week"][z] = 0
    for m in mm:
        days_dict[act]["month"][m] = 0

print("开始制作'52 Weeks'表单\n")
# 整理数据
for row in ws2.iter_rows(min_row=2, max_row=ws2.max_row):
    tt = row[0].value
    z_tt = tt.strftime("%W")
    m_tt = str(tt.month)
    for act in activity_name:
        num = activity_name.index(act)+1  # 列表的索引与实际引用的表格索引不同
        days_dict[act]["week"][z_tt] += row[num].value
        days_dict[act]["month"][m_tt] += row[num].value
# 填写数据
for z in zz:
    i = ["第"+str(int(z)+1)+"周"]
    for act in activity_name:
        i.append(round(days_dict[act]["week"][z], 1))
    ws3.append(i)
print("'52 Weeks'表单制作完成\n")
"""
表单4:"12 Months"
"""
# 填写数据
for m in mm:
    i = ["第"+m+"月"]
    for act in activity_name:
        i.append(round(days_dict[act]["month"][m], 1))
    ws4.append(i)
print("'12 Months'表单制作完成\n")

# 保存
wb.save(save_url)
