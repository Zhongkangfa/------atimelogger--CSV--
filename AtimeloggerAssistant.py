import csv
import openpyxl
import datetime
from Activity import *


class AtimeloggerAssistant:
    def __init__(self):
        self.raw_intervals = []
        self.raw_types = []
        self.types = set()
        self.group_types = set()
        self.no_group_types = set()
        self.activitys = []
        pass

    def _create_sheet(self):
        self.wb = openpyxl.Workbook()
        self.ws_days = self.wb.active
        self.ws_days.title = "365 Days"
        self.ws_weeks = self.wb.create_sheet("52 Weeks")
        self.ws_months = self.wb.create_sheet("12 Months")

    def add_report(self, csv_path):
        self.file_path = csv_path
        pass

    def work(self):
        self._separate()
        self._get_activity_type_info()
        self._create_activitys()
        pass


    def summarizing(self):
        self._create_sheet()
        table_header = list(self.no_group_types)
        table_header.insert(0, "")
        self.ws_days.append(table_header)
        self.ws_weeks.append(table_header)
        self.ws_months.append(table_header)
        days = sorted(list(set([datetime.datetime.strptime(
            interval[2], "%Y/%m/%d %H:%M:%S").date() for interval in self.raw_intervals])))

        # 制作第一张表单
        for day in days:
            row = [str(day)]
            for activity in self.activitys:
                counted_activity_time = activity.total_by_day(day)
                row.append(counted_activity_time)
            self.ws_days.append(row)

        # 第二张
        weeks = sorted(list(set([day.strftime("%Y-%W") for day in days])))
        for week in weeks:
            row = [week]
            # row = [str(life_date.year)+"-"+str(life_date.month)]
            for activity in self.activitys:
                counted_activity_time = activity.total_by_week(week)
                row.append(counted_activity_time)
            self.ws_weeks.append(row)

        # 第三张
        months = sorted(list(set([day.strftime("%Y-%m") for day in days])))
        for month in months:
            row = [month]
            for activity in self.activitys:
                counted_activity_time = activity.total_by_month(month)
                row.append(counted_activity_time)
            self.ws_months.append(row)

    def save(self, file_name):
        if ".xlse" not in file_name:
            file_name = file_name + ".xlsx"
        self.wb.save(file_name)

    def _separate(self):
        """
        导出来的CSV文件中，通过与时间记录空一行的形式汇总了活动类别，保存着活动类别的层级关系。
        第一步就是要将这两个部分区分出来。

        """
        with open(self.file_path, 'r', encoding='utf-8') as f:
            read = csv.reader(f)
            read.__next__()

            # 收集原始的intervals，待进一步处理
            for row in read:
                if row:
                    self.raw_intervals.append(row[-5:])
                elif not row:
                    break

            read.__next__()

            # 收集原始的types
            for row in read:
                self.raw_types.append(row)
        pass

    def _get_activity_type_info(self):
        """
        我的程序暂不支持重名活动类别
        主要是区分出活动组（Group）和活动
        先统计出具体活动的情况，才方便再进一步去汇总分组情况。
        """
        # all_types = group + activity
        self.all_types = set([raw_type[-3] for raw_type in self.raw_types])
        # group_types = group
        self.group_types = set([raw_type[-4]
                                for raw_type in self.raw_types if raw_type[-4]])
        # no_group_types = activity
        self.no_group_types = self.all_types ^ self.group_types

        pass

    def _create_activitys(self):
        for activity_name in self.no_group_types:
            own_intervals = [
                interval for interval in self.raw_intervals if interval[0] == activity_name]
            # ['形象', '00:32:35', '2019/5/5 21:59:33', '2019/5/5 22:32:08', '']
            self.activitys.append(Activity(activity_name, own_intervals))
        self.activitys = sorted(self.activitys)
        pass
